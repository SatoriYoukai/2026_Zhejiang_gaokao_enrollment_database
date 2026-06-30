from __future__ import annotations

import argparse
import bisect
import csv
import math
import random
import re
import sys
from collections import defaultdict
from pathlib import Path


POPULATION_DEFAULT = 400_000
USER_RANK_DEFAULT = 39_000


def resource_path(relative: str) -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parents[1]))
    return base / relative


def default_master_path() -> Path:
    bundled = resource_path("data/rank_model_database.csv")
    if bundled.exists():
        return bundled
    return Path.home() / "Documents" / "志愿填报" / "outputs" / "clean_database" / "volunteer_master_2026_with_history.csv"


def norm_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def safe_float(value: object) -> float | None:
    text = norm_text(value)
    if not text:
        return None
    try:
        value = float(text)
    except ValueError:
        return None
    if math.isnan(value):
        return None
    return value


def source_label(value: object) -> str:
    labels = {
        "2025_history": "2025年历史位次",
        "2024_history": "2024年历史位次",
        "2023_history": "2023年历史位次",
        "same_school_all_majors_trimmed_mean": "无历史：同校所有专业截尾均值",
        "missing_need_user_input": "缺失：需要用户输入",
        "user_manual_input": "用户手动输入",
    }
    return labels.get(norm_text(value), norm_text(value))


def confidence_label(value: object) -> str:
    labels = {
        "high": "高",
        "medium_high": "中高",
        "low": "低",
        "need_user_input": "需要用户输入",
        "manual": "手动输入",
    }
    return labels.get(norm_text(value), norm_text(value))


def normalize_code(value: object, width: int) -> str:
    text = norm_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    return digits.zfill(width) if digits else text


def logit_rank(rank: float, population: int) -> float:
    p = min(max(rank / population, 1e-6), 1 - 1e-6)
    return math.log(p / (1 - p))


def major_family(major: str) -> str:
    major = norm_text(major)
    if any(k in major for k in ["计算机", "软件工程", "网络工程"]):
        return "cs"
    if any(k in major for k in ["人工智能", "智能科学"]):
        return "ai"
    if "信息与计算科学" in major:
        return "info_compute"
    if any(k in major for k in ["统计", "应用统计"]):
        return "statistics"
    if any(k in major for k in ["数学", "数据计算"]):
        return "math"
    return "other"


def rank_bin_label(rank: float) -> str:
    bins = [
        (0, 20_000, "0_20k"),
        (20_000, 30_000, "20k_30k"),
        (30_000, 37_000, "30k_37k"),
        (37_000, 40_000, "37k_40k"),
        (40_000, 50_000, "40k_50k"),
        (50_000, 60_000, "50k_60k"),
        (60_000, 75_000, "60k_75k"),
        (75_000, 100_000, "75k_100k"),
        (100_000, 150_000, "100k_150k"),
        (150_000, 250_000, "150k_250k"),
        (250_000, 400_000, "250k_400k"),
    ]
    for lo, hi, label in bins:
        if lo <= rank <= hi:
            return label
    return "out"


def plan_bin(plan: float | None) -> str:
    if plan is None:
        return "unknown"
    if plan <= 2:
        return "plan_1_2"
    if plan <= 5:
        return "plan_3_5"
    if plan <= 10:
        return "plan_6_10"
    if plan <= 30:
        return "plan_11_30"
    return "plan_31_plus"


def is_zhejiang(province: str) -> bool:
    return norm_text(province) == "浙江"


def is_hat_school(college: str) -> bool:
    return any(k in norm_text(college) for k in ["双一流", "985", "211"])


def read_csv_rows(path: Path) -> list[list[object]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return [row for row in csv.reader(f)]


def read_xlsx_rows(path: Path) -> list[list[object]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def read_xls_rows(path: Path) -> list[list[object]]:
    import xlrd

    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)
    return [ws.row_values(i) for i in range(ws.nrows)]


def row_looks_like_header(row: list[object]) -> bool:
    text = "".join(norm_text(x) for x in row[:5])
    if any(k in text for k in ["院校", "学校", "专业", "代码", "college", "major"]):
        return True
    first = norm_text(row[0]) if row else ""
    return not bool(re.fullmatch(r"\d{4}(\.0)?", first))


def read_volunteer_input(path: Path) -> list[dict]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        rows = read_csv_rows(path)
    elif suffix == ".xlsx":
        rows = read_xlsx_rows(path)
    elif suffix == ".xls":
        rows = read_xls_rows(path)
    else:
        raise ValueError(f"不支持的文件类型: {suffix}")

    rows = [row for row in rows if any(norm_text(x) for x in row)]
    if not rows:
        raise ValueError("志愿表为空")

    header_map = {
        "院校代码": "college_code",
        "学校代码": "college_code",
        "college_code": "college_code",
        "院校名称": "college",
        "学校名称": "college",
        "college": "college",
        "专业代码": "major_code",
        "major_code": "major_code",
        "专业名称": "major",
        "major": "major",
        "志愿序号": "order",
        "序号": "order",
        "order": "order",
    }
    volunteers: list[dict] = []

    if row_looks_like_header(rows[0]):
        header = [norm_text(x) for x in rows[0]]
        col_to_index = {}
        for idx, col in enumerate(header):
            if col in header_map:
                col_to_index[header_map[col]] = idx
        if {"college_code", "college", "major_code", "major"}.issubset(col_to_index):
            data_rows = rows[1:]
            for i, row in enumerate(data_rows, start=1):
                if len(row) <= max(col_to_index.values()):
                    continue
                order_idx = col_to_index.get("order")
                order_value = safe_float(row[order_idx]) if order_idx is not None else None
                volunteers.append(
                    {
                        "order": int(order_value) if order_value else i,
                        "college_code": normalize_code(row[col_to_index["college_code"]], 4),
                        "college": norm_text(row[col_to_index["college"]]),
                        "major_code": normalize_code(row[col_to_index["major_code"]], 3),
                        "major": norm_text(row[col_to_index["major"]]),
                    }
                )
            return finalize_volunteers(volunteers)
        rows = rows[1:]

    for i, row in enumerate(rows, start=1):
        if len(row) < 4:
            continue
        volunteers.append(
            {
                "order": i,
                "college_code": normalize_code(row[0], 4),
                "college": norm_text(row[1]),
                "major_code": normalize_code(row[2], 3),
                "major": norm_text(row[3]),
            }
        )
    return finalize_volunteers(volunteers)


def finalize_volunteers(volunteers: list[dict]) -> list[dict]:
    cleaned = []
    for item in volunteers:
        if not any([item["college_code"], item["college"], item["major_code"], item["major"]]):
            continue
        item["volunteer_id"] = f'{item["college_code"]}-{item["major_code"]}'
        cleaned.append(item)
    cleaned.sort(key=lambda x: int(x["order"]))
    for idx, item in enumerate(cleaned, start=1):
        item["order"] = idx
    return cleaned


def load_master(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    for row in rows:
        row["college"] = row.get("college_name") or row.get("college") or ""
        row["major"] = row.get("major_name") or row.get("major") or ""
        row["plan_count_num"] = safe_float(row.get("plan_count"))
        row["rank_ref_num"] = safe_float(row.get("rank_ref"))
        row["rank_ref_source_label"] = source_label(row.get("rank_ref_source"))
        row["rank_ref_confidence_label"] = confidence_label(row.get("rank_ref_confidence"))
        row["major_family"] = row.get("major_family_model") or major_family(row["major"])
        row["rank_bin"] = rank_bin_label(row["rank_ref_num"]) if row["rank_ref_num"] else "unknown"
        row["plan_bin"] = plan_bin(row["plan_count_num"])
        row["is_zhejiang"] = is_zhejiang(row.get("province", ""))
        row["is_hat"] = is_hat_school(row["college"])
    return rows


def median(values: list[float]) -> float:
    if not values:
        return 0.0
    xs = sorted(values)
    mid = len(xs) // 2
    if len(xs) % 2:
        return xs[mid]
    return (xs[mid - 1] + xs[mid]) / 2


def build_transition_samples(master: list[dict], population: int, trend_mode: str) -> list[dict]:
    samples: list[dict] = []
    pairs = [
        ("2023_history_lowest_rank", "2024_history_lowest_rank"),
        ("2024_history_lowest_rank", "2025_history_lowest_rank"),
    ]
    for row in master:
        for c0, c1 in pairs:
            old = safe_float(row.get(c0))
            new = safe_float(row.get(c1))
            if old is None or new is None or old <= 0 or new <= 0:
                continue
            rank_mid = (old + new) / 2
            samples.append(
                {
                    "rank_mid": rank_mid,
                    "delta_z": logit_rank(new, population) - logit_rank(old, population),
                    "major_family": row["major_family"],
                    "rank_bin": rank_bin_label(rank_mid),
                    "plan_bin": row["plan_bin"],
                    "plan_count": row["plan_count_num"],
                    "is_zhejiang": row["is_zhejiang"],
                    "is_hat": row["is_hat"],
                }
            )
    if not samples:
        raise ValueError("内置数据库中没有可用的 2023-2025 历史位次样本")

    if trend_mode == "raw":
        for sample in samples:
            sample["model_delta_z"] = sample["delta_z"]
    elif trend_mode == "global_centered":
        center = median([s["delta_z"] for s in samples])
        for sample in samples:
            sample["model_delta_z"] = sample["delta_z"] - center
    elif trend_mode == "rankbin_centered":
        grouped: dict[str, list[float]] = defaultdict(list)
        for sample in samples:
            grouped[sample["rank_bin"]].append(sample["delta_z"])
        centers = {key: median(values) for key, values in grouped.items()}
        for sample in samples:
            sample["model_delta_z"] = sample["delta_z"] - centers.get(sample["rank_bin"], 0.0)
    else:
        raise ValueError(f"未知 trend-mode: {trend_mode}")
    return samples


def parse_rank_input(text: str) -> float | None:
    text = text.strip().replace(",", "").replace("，", "")
    return safe_float(text)


def ask_user_rank_for_missing(item: dict, population: int, reason: str) -> float:
    print("")
    print("需要手动输入参考位次")
    print(f'志愿 {int(item["order"]):02d}: {item["volunteer_id"]} {item["college"]} {item["major"]}')
    print(reason)
    while True:
        text = input(f"请输入该志愿的参考位次（1-{population}，直接回车则停止）：").strip()
        if not text:
            raise ValueError("用户未输入缺失位次，已停止估算。")
        rank = parse_rank_input(text)
        if rank is not None and 1 <= rank <= population:
            return rank
        print("位次格式不正确，请输入类似 39000 的数字。")


def ask_user_rank(default: float = USER_RANK_DEFAULT) -> float:
    print(f"请输入考生位次，直接回车使用默认 {default:.0f}：")
    while True:
        text = input().strip()
        if not text:
            return default
        rank = parse_rank_input(text)
        if rank is not None and rank > 0:
            return rank
        print("位次格式不正确，请输入类似 39000 的数字：")


def ask_input_path() -> Path:
    print("")
    print("请把志愿录入表文件拖到这个窗口，或粘贴完整路径，然后回车：")
    while True:
        text = input().strip().strip('"')
        if text:
            path = Path(text)
            if path.exists():
                return path
            print("输入文件不存在，请重新拖入或粘贴路径：")
        else:
            print("路径不能为空，请重新输入：")


def confirm_predicted_ranks(rows: list[dict], policy: str) -> bool:
    if not rows:
        return True
    if policy == "accept":
        return True
    if policy == "manual":
        return False

    print("")
    print("以下志愿完全无历史位次，程序使用“同校所有专业截尾均值”生成了预测参考位次：")
    print("")
    for row in rows:
        print(f'{int(row["order"]):02d}. {row["volunteer_id"]} {row["college"]} {row["major"]} -> 预测参考位次 {row["rank_ref_num"]:.0f}')
    print("")
    print("这类位次只用于结构估算，不是真实历史录取位次。")
    while True:
        answer = input("是否接受以上预测位次？输入 y 接受；输入 n 则逐条手动输入这些志愿的参考位次：").strip().lower()
        if answer in {"y", "yes", "是", "接受"}:
            return True
        if answer in {"n", "no", "否", "不接受"}:
            return False
        print("请输入 y 或 n。")


def set_manual_rank(row: dict, rank_ref_num: float) -> None:
    row["rank_ref_num"] = rank_ref_num
    row["rank_ref"] = rank_ref_num
    row["rank_ref_source"] = "user_manual_input"
    row["rank_ref_confidence"] = "manual"
    row["rank_ref_source_label"] = source_label(row["rank_ref_source"])
    row["rank_ref_confidence_label"] = confidence_label(row["rank_ref_confidence"])
    row["rank_bin_model"] = rank_bin_label(row["rank_ref_num"])


def merge_master_info(volunteers: list[dict], master: list[dict], population: int, predicted_rank_policy: str) -> list[dict]:
    by_id = {row["volunteer_id"]: row for row in master if row.get("volunteer_id")}
    manual = 0
    merged = []
    predicted_rows = []
    missing_rows = []
    for item in volunteers:
        master_row = by_id.get(item["volunteer_id"])
        if not master_row:
            master_row = {}
            rank_ref_num = None
            missing_reason = "内置数据库中没有找到这条志愿。"
        else:
            rank_ref_num = float(master_row["rank_ref_num"]) if master_row.get("rank_ref_num") is not None else None
            missing_reason = "内置数据库中有这条志愿，但没有历史位次，也无法用同校截尾均值预测。"
        row = dict(item)
        for key in [
            "province",
            "city",
            "plan_count",
            "tuition",
            "subject_requirement",
            "remark",
            "rank_ref",
            "rank_ref_source",
            "rank_ref_confidence",
            "2023_history_lowest_rank",
            "2024_history_lowest_rank",
            "2025_history_lowest_rank",
        ]:
            row[key] = master_row.get(key, "")
        if master_row:
            row["rank_ref_num"] = rank_ref_num
            row["plan_count_num"] = master_row["plan_count_num"]
        else:
            row["rank_ref_num"] = rank_ref_num
            row["plan_count_num"] = None
        row["rank_ref_source_label"] = source_label(row.get("rank_ref_source"))
        row["rank_ref_confidence_label"] = confidence_label(row.get("rank_ref_confidence"))
        row["major_family_model"] = major_family(row["major"])
        row["rank_bin_model"] = rank_bin_label(row["rank_ref_num"]) if row["rank_ref_num"] is not None else "unknown"
        row["plan_bin_model"] = plan_bin(row["plan_count_num"])
        row["is_zhejiang_model"] = is_zhejiang(row.get("province", ""))
        row["is_hat_model"] = is_hat_school(row["college"])
        history_count = sum(1 for c in ["2023_history_lowest_rank", "2024_history_lowest_rank", "2025_history_lowest_rank"] if safe_float(row.get(c)) is not None)
        row["history_count_model"] = history_count
        if row.get("rank_ref_source") == "same_school_all_majors_trimmed_mean":
            predicted_rows.append(row)
        if row["rank_ref_num"] is None:
            row["missing_rank_reason"] = missing_reason
            missing_rows.append(row)
        merged.append(row)
    if predicted_rows:
        accepted = confirm_predicted_ranks(predicted_rows, predicted_rank_policy)
        if not accepted:
            for row in predicted_rows:
                rank_ref_num = ask_user_rank_for_missing(
                    row,
                    population,
                    f'当前预测参考位次为 {row["rank_ref_num"]:.0f}，来源是同校所有专业截尾均值；你选择不接受，需要手动输入。',
                )
                set_manual_rank(row, rank_ref_num)
                manual += 1
    for row in missing_rows:
        rank_ref_num = ask_user_rank_for_missing(row, population, row["missing_rank_reason"])
        set_manual_rank(row, rank_ref_num)
        manual += 1
    estimated = sum(
        1
        for row in merged
        if row.get("rank_ref_source") == "same_school_all_majors_trimmed_mean"
        or row.get("rank_ref_source") == "user_manual_input"
    )
    if estimated:
        print(f"提示：{estimated} 条志愿使用预测/手填位次，而不是直接历史位次。")
    if manual:
        print(f"提示：其中 {manual} 条志愿的参考位次来自用户手动输入。")
    return merged


def filter_samples(samples: list[dict], cand: dict) -> tuple[list[dict], str]:
    rank = cand["rank_ref_num"]
    near_lo = max(1, rank * 0.65)
    near_hi = min(399_999, rank * 1.45)

    def same_rank(s: dict) -> bool:
        return near_lo <= s["rank_mid"] <= near_hi

    masks = [
        (lambda s: same_rank(s) and s["major_family"] == cand["major_family_model"] and s["plan_bin"] == cand["plan_bin_model"] and s["is_zhejiang"] == cand["is_zhejiang_model"], "rank+family+plan+region"),
        (lambda s: same_rank(s) and s["major_family"] == cand["major_family_model"] and s["plan_bin"] == cand["plan_bin_model"], "rank+family+plan"),
        (lambda s: same_rank(s) and s["major_family"] == cand["major_family_model"], "rank+family"),
        (lambda s: s["rank_bin"] == cand["rank_bin_model"] and s["major_family"] == cand["major_family_model"], "rankbin+family"),
        (same_rank, "near_rank"),
        (lambda s: s["rank_bin"] == cand["rank_bin_model"], "rankbin"),
    ]
    for predicate, label in masks:
        subset = [s for s in samples if predicate(s)]
        if len(subset) >= 80:
            return subset, label
    subset = [s for s in samples if same_rank(s)]
    return (subset if subset else samples), "near_rank_fallback"


def weighted_pool(samples: list[dict], cand: dict, population: int) -> tuple[list[float], list[float], dict]:
    subset, tier = filter_samples(samples, cand)
    rank_z = logit_rank(cand["rank_ref_num"], population)
    plan_value = cand["plan_count_num"] if cand["plan_count_num"] is not None else 3.0
    deltas = []
    weights = []
    for sample in subset:
        z_dist = abs(logit_rank(sample["rank_mid"], population) - rank_z)
        sample_plan = sample["plan_count"] if sample["plan_count"] is not None else plan_value
        plan_dist = abs(math.log1p(sample_plan) - math.log1p(max(0.0, plan_value)))
        family_bonus = 0.75 if sample["major_family"] == cand["major_family_model"] else 1.0
        region_bonus = 0.85 if sample["is_zhejiang"] == cand["is_zhejiang_model"] else 1.0
        weight = math.exp(-z_dist / 0.22) * math.exp(-plan_dist / 1.2) * family_bonus * region_bonus
        if weight > 0:
            deltas.append(sample["model_delta_z"])
            weights.append(weight)

    if len(deltas) < 40:
        fallback = [s for s in samples if s["rank_bin"] == cand["rank_bin_model"]] or samples
        local_weight = 0.55 if len(deltas) >= 15 else 0.35
        total = sum(weights)
        if total > 0:
            weights = [w / total * local_weight for w in weights]
        fallback_weight = (1 - local_weight) / len(fallback)
        deltas.extend(s["model_delta_z"] for s in fallback)
        weights.extend([fallback_weight] * len(fallback))
        tier = f"{tier}+rankbin_shrink"

    total = sum(weights)
    weights = [w / total for w in weights]
    cumulative = []
    acc = 0.0
    for weight in weights:
        acc += weight
        cumulative.append(acc)
    cumulative[-1] = 1.0
    eff_n = 1.0 / sum(w * w for w in weights)
    return deltas, cumulative, {"sample_tier": tier, "sample_n": len(deltas), "effective_n": eff_n}


def sample_from_pool(deltas: list[float], cumulative: list[float], rng: random.Random) -> float:
    idx = bisect.bisect_left(cumulative, rng.random())
    if idx >= len(deltas):
        idx = len(deltas) - 1
    return deltas[idx]


def model_note(row: dict) -> str:
    history_count = int(row["history_count_model"])
    if history_count >= 3:
        parts = ["3年历史"]
    elif history_count == 2:
        parts = ["2年历史"]
    elif history_count == 1:
        parts = ["1年历史"]
    else:
        parts = ["无历史"]
    if row["plan_count_num"] is not None and row["plan_count_num"] <= 3:
        parts.append("小计划")
    if row.get("rank_ref_source") and row["rank_ref_source"] not in {"2025_history", "2024_history", "2023_history"}:
        parts.append(source_label(row["rank_ref_source"]))
    parts.append(str(row["sample_tier"]))
    parts.append(f'eff_n={float(row["effective_n"]):.0f}')
    return ";".join(parts)


def simulate(volunteers: list[dict], samples: list[dict], user_rank: float, population: int, draws: int, seed: int) -> list[dict]:
    rng = random.Random(seed)
    user_z = logit_rank(user_rank, population)
    params = []
    for cand in volunteers:
        deltas, cumulative, info = weighted_pool(samples, cand, population)
        history_count = int(cand["history_count_model"])
        extra_sigma = 0.0
        if history_count == 0:
            extra_sigma += 0.055
        elif history_count == 1:
            extra_sigma += 0.035
        if cand["plan_count_num"] is not None and cand["plan_count_num"] <= 3:
            extra_sigma += 0.020
        cand.update(info)
        params.append((logit_rank(cand["rank_ref_num"], population), deltas, cumulative, extra_sigma))

    n = len(volunteers)
    can_counts = [0] * n
    admit_counts = [0] * n
    sqrt2 = math.sqrt(2)
    for _ in range(draws):
        first = None
        for j, (base_z, deltas, cumulative, extra_sigma) in enumerate(params):
            shock = sample_from_pool(deltas, cumulative, rng) / sqrt2
            if extra_sigma:
                shock += rng.gauss(0.0, extra_sigma)
            can = base_z + shock >= user_z
            if can:
                can_counts[j] += 1
                if first is None:
                    first = j
        if first is not None:
            admit_counts[first] += 1

    admitted_before = 0
    cumulative_admit = 0.0
    recursive_reach = 1.0
    recursive_cumulative = 0.0
    result = []
    for j, row in enumerate(volunteers):
        reach_count = draws - admitted_before
        standalone_can_prob = can_counts[j] / draws
        recursive_admit_prob = recursive_reach * standalone_can_prob
        recursive_cumulative += recursive_admit_prob
        admit_prob = admit_counts[j] / draws
        cumulative_admit += admit_prob
        out = dict(row)
        out["reach_prob"] = reach_count / draws
        out["prob_if_reached"] = admit_counts[j] / reach_count if reach_count else 0.0
        out["standalone_can_prob"] = standalone_can_prob
        out["recursive_reach_prob"] = recursive_reach
        out["recursive_admit_prob"] = recursive_admit_prob
        out["recursive_cum_admit_prob"] = recursive_cumulative
        out["admit_prob"] = admit_prob
        out["cum_admit_prob"] = cumulative_admit
        out["no_admit_after"] = max(0.0, 1 - cumulative_admit)
        out["recursive_no_admit_after"] = max(0.0, 1 - recursive_cumulative)
        out["rank_gap"] = out["rank_ref_num"] - user_rank
        out["model_uncertainty_note"] = model_note(out)
        result.append(out)
        admitted_before += admit_counts[j]
        recursive_reach *= (1 - standalone_can_prob)
    return result


def pct(value: float) -> str:
    return f"{value * 100:.2f}%"


OUTPUT_COLUMNS = [
    ("order", "志愿序号"),
    ("volunteer_id", "志愿代码"),
    ("college_code", "院校代码"),
    ("college", "院校名称"),
    ("major_code", "专业代码"),
    ("major", "专业名称"),
    ("province", "省份"),
    ("city", "城市"),
    ("plan_count", "招生计划数"),
    ("tuition", "学费"),
    ("subject_requirement", "选科要求"),
    ("remark", "备注"),
    ("rank_ref_num", "参考位次"),
    ("rank_ref_source_label", "参考位次来源"),
    ("rank_ref_confidence_label", "参考位次置信度"),
    ("2023_history_lowest_rank", "2023最低位次"),
    ("2024_history_lowest_rank", "2024最低位次"),
    ("2025_history_lowest_rank", "2025最低位次"),
    ("rank_gap", "参考位次差"),
    ("reach_prob", "到达概率"),
    ("prob_if_reached", "到达后可录概率"),
    ("standalone_can_prob", "单独可录概率"),
    ("recursive_reach_prob", "递推到达概率"),
    ("recursive_admit_prob", "递推落点概率"),
    ("recursive_cum_admit_prob", "递推累计录取概率"),
    ("recursive_no_admit_after", "递推该志愿后仍未录取概率"),
    ("admit_prob", "最终落点概率"),
    ("cum_admit_prob", "累计录取概率"),
    ("no_admit_after", "该志愿后仍未录取概率"),
    ("model_uncertainty_note", "模型备注"),
]


def output_value(row: dict, key: str) -> object:
    value = row.get(key, "")
    if key in {
        "reach_prob",
        "prob_if_reached",
        "standalone_can_prob",
        "recursive_reach_prob",
        "recursive_admit_prob",
        "recursive_cum_admit_prob",
        "recursive_no_admit_after",
        "admit_prob",
        "cum_admit_prob",
        "no_admit_after",
    }:
        return pct(float(value)) if value != "" else ""
    if key in {"rank_ref_num", "rank_gap", "2023_history_lowest_rank", "2024_history_lowest_rank", "2025_history_lowest_rank"}:
        num = safe_float(value)
        return round(num, 2) if num is not None and abs(num - round(num)) > 0.005 else (int(round(num)) if num is not None else "")
    if key == "plan_count":
        num = safe_float(value)
        return int(round(num)) if num is not None and abs(num - round(num)) < 0.005 else (value or "")
    return value


def write_csv(path: Path, rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow([label for _, label in OUTPUT_COLUMNS])
        for row in rows:
            writer.writerow([output_value(row, key) for key, _ in OUTPUT_COLUMNS])


def write_xlsx(path: Path, rows: list[dict]) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "落点概率"
    ws.append([label for _, label in OUTPUT_COLUMNS])
    for row in rows:
        ws.append([output_value(row, key) for key, _ in OUTPUT_COLUMNS])
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 48)
    wb.save(path)


def write_report(path: Path, input_path: Path, rows: list[dict], user_rank: float, population: int, trend_mode: str, draws: int) -> None:
    total = sum(row["admit_prob"] for row in rows)
    source_counts: dict[str, int] = defaultdict(int)
    for row in rows:
        source_counts[row.get("rank_ref_source_label", "未知")] += 1
    lines = [
        "志愿落点概率估算报告",
        "=" * 24,
        "",
        f"输入志愿表：{input_path.name}",
        f"考生位次：{user_rank:.0f}",
        f"全省参考人数：{population}",
        f"趋势处理：{trend_mode}",
        f"模拟次数：{draws}",
        "",
        "概览",
        "-" * 24,
        "",
        f"志愿表内录取概率估计：{pct(total)}",
        f"志愿表后仍未录取概率估计：{pct(max(0.0, 1 - total))}",
        "",
        "参考位次来源",
        "-" * 24,
        "",
    ]
    for label, count in sorted(source_counts.items(), key=lambda x: (-x[1], x[0])):
        lines.append(f"{label}：{count} 条")
    lines.extend([
        "",
        "最可能落点",
        "-" * 24,
        "",
    ])
    for row in sorted(rows, key=lambda x: x["admit_prob"], reverse=True)[:20]:
        lines.append(
            f'{int(row["order"]):02d}. {row["college"]} {row["major"]} | 参考位次 {row["rank_ref_num"]:.0f} | '
            f'到达 {pct(row["reach_prob"])} | 到达后可录 {pct(row["prob_if_reached"])} | '
            f'最终落点 {pct(row["admit_prob"])} | {row["model_uncertainty_note"]}'
        )
    lines.extend(["", "十志愿窗口", "-" * 24, ""])
    max_order = max(int(row["order"]) for row in rows)
    by_order = {int(row["order"]): row for row in rows}
    for lo in range(1, max_order + 1, 10):
        hi = min(lo + 9, max_order)
        chunk = [by_order[i] for i in range(lo, hi + 1) if i in by_order]
        lines.append(
            f'{lo:02d}-{hi:02d}: 落点合计 {pct(sum(r["admit_prob"] for r in chunk))}, '
            f'窗口前到达 {pct(chunk[0]["reach_prob"])}, '
            f'窗口后未录 {pct(chunk[-1]["no_admit_after"])}'
        )
    lines.extend(
        [
            "",
            "模型口径简述",
            "-" * 24,
            "",
            "1. 有历史位次的志愿使用最近一年历史最低录取位次；2025 优先，其次 2024，再其次 2023。",
            "2. 完全无历史位次的志愿，使用同校所有有历史专业的截尾均值作为参考位次；如果同校也无法提供锚点，则运行时要求用户手动输入。",
            "3. 概率模型使用 2023->2024、2024->2025 的年际位次变化做经验抽样，再按浙江专业平行志愿规则模拟检索。",
            "4. 这个结果适合检查志愿表结构，不适合把小数点后的概率当作精确预测。",
        ]
    )
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Estimate Zhejiang professional-parallel volunteer landing probabilities.")
    parser.add_argument("input", type=Path, nargs="?", help="Volunteer table: .xls/.xlsx/.csv")
    parser.add_argument("--master", type=Path, default=default_master_path(), help="Bundled rank model database")
    parser.add_argument("--out-dir", type=Path, default=None, help="Output directory; default is input file directory")
    parser.add_argument("--user-rank", type=float, default=None)
    parser.add_argument("--population", type=int, default=POPULATION_DEFAULT)
    parser.add_argument("--draws", type=int, default=200_000)
    parser.add_argument("--seed", type=int, default=20260630)
    parser.add_argument("--trend-mode", choices=["raw", "global_centered", "rankbin_centered"], default="rankbin_centered")
    parser.add_argument(
        "--predicted-rank-policy",
        choices=["ask", "accept", "manual"],
        default="ask",
        help="How to handle no-history ranks estimated by same-school trimmed mean.",
    )
    args = parser.parse_args()

    if args.user_rank is None:
        args.user_rank = ask_user_rank()
    if args.input is None:
        args.input = ask_input_path()

    out_dir = args.out_dir or args.input.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    volunteers = read_volunteer_input(args.input)
    master = load_master(args.master)
    samples = build_transition_samples(master, args.population, args.trend_mode)
    merged = merge_master_info(volunteers, master, args.population, args.predicted_rank_policy)
    result = simulate(merged, samples, args.user_rank, args.population, args.draws, args.seed)

    stem = args.input.stem
    csv_path = out_dir / f"{stem}_落点概率估算.csv"
    xlsx_path = out_dir / f"{stem}_落点概率估算.xlsx"
    report_path = out_dir / f"{stem}_落点概率估算报告.txt"
    write_csv(csv_path, result)
    write_xlsx(xlsx_path, result)
    write_report(report_path, args.input, result, args.user_rank, args.population, args.trend_mode, args.draws)

    print("输入文件", args.input)
    print("志愿数量", len(result))
    print("CSV结果", csv_path)
    print("Excel结果", xlsx_path)
    print("TXT报告", report_path)
    print("最可能落点")
    for row in sorted(result, key=lambda x: x["admit_prob"], reverse=True)[:12]:
        print(
            f'{int(row["order"]):02d} {row["college"]} {row["major"]} '
            f'参考位次={row["rank_ref_num"]:.0f} 到达={pct(row["reach_prob"])} '
            f'到达后可录={pct(row["prob_if_reached"])} 落点={pct(row["admit_prob"])}'
        )


if __name__ == "__main__":
    main()
